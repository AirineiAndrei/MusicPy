from __future__ import unicode_literals
import os
import openpyxl
import yt_dlp

# Define default path
path = r"D:/MusicPy/music"

# Define download rate limit in byte
ratelimit = 50000000

# Define download format
format = 'ba' # Best possible audio



def DownloadYoutube(url):
    # Download all videos in a playlist
    if url.startswith('https://www.youtube.com/playlist'):
        ydl_opts = {
            'abort_on_unavailable_fragments': True,
            'format': format,
            'outtmpl': path + '\\%(playlist_uploader)s ## %(playlist)s\%(title)s.%(ext)s',
            'ratelimit': ratelimit,
        }

    # Download single video from url
    elif url.startswith((
        'https://www.youtube.com/watch')):
        ydl_opts = {
            'abort_on_unavailable_fragments': True,
            'format': format,
            'outtmpl': path + '\\%(title)s ## %(uploader)s.%(ext)s',
            'ratelimit': ratelimit,
        }

    # Downloads depending on the options set above
    if ydl_opts is not None:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            ydl.download(url)
    


script_dir = os.path.dirname(__file__)
rel_path = "data/test.xlsx"
abs_file_path = os.path.join(script_dir, rel_path)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(abs_file_path)

# Define variable to read sheet
dataframe1 = dataframe.active

HadError = False
# Iterate the read links in cells
with open('logs.txt', 'w') as logout:
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            link = col[row].value
            try:
                if "youtube" in link: # youtube link
                    DownloadYoutube(link)
            except:
                HadError = True
                logout.write(f"Failed to Download: {link} \n")

if HadError:
    print("An error occured, check log files")
